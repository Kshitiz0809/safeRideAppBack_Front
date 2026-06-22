import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h
def fill(hex_): return PatternFill("solid", fgColor=hex_)
def fnt(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

INR_FMT = u'"₹"#,##0'

NAVY, TEAL, GREEN, AMBER, RED = "1B2A4A", "0D7377", "27AE60", "F39C12", "E74C3C"
LGRAY, MGRAY, WHITE, DGRAY   = "F4F6F8", "DDE1E7", "FFFFFF", "5A6475"

def hdr(ws, row, vals, bg=NAVY, fg=WHITE, sz=11):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(row, c, v)
        cl.fill, cl.font, cl.alignment, cl.border = fill(bg), fnt(True, sz, fg), aln("center"), bdr()

def drow(ws, row, vals, bg=WHITE, bold=False, fg="1B2A4A"):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(row, c, v)
        cl.fill, cl.font, cl.alignment, cl.border = fill(bg), fnt(bold, 10, fg), aln("left"), bdr()

# SHEET 1 SUMMARY
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
for i, w in enumerate([3, 34, 30, 18, 18, 3], 1): cw(ws1, i, w)

ws1.merge_cells("B1:E1")
cl = ws1["B1"]; cl.value = "SafeRider App - Cost & CI/CD Budget Breakdown"
cl.fill, cl.font, cl.alignment = fill(NAVY), fnt(True, 18, WHITE), aln("center")
rh(ws1, 1, 40)

ws1.merge_cells("B2:E2")
cl = ws1["B2"]; cl.value = "All amounts in Indian Rupees (Rs)  |  Prepared 2026"
cl.fill, cl.font, cl.alignment = fill(TEAL), fnt(italic=True, size=10, color=WHITE), aln("center")
rh(ws1, 2, 18)
rh(ws1, 3, 8)

hdr(ws1, 4, ["Phase", "Description", "Monthly (Rs)", "Annual (Rs)"], bg=TEAL)
rh(ws1, 4, 22)

rows = [
    ("Phase 0  Free",    "HF Spaces sleep, Firebase free tier, GitHub free",           0,       0),
    ("Phase 1  Starter", "Up to 1K users: Railway Rs420/mo, Firebase free",          420,    5040),
    ("Phase 2  Growth",  "1K-10K users: Firebase paid avg + Railway",               2940,   35280),
    ("Phase 3  Scale",   "10K-100K users: Cloud infra upgrade",                    12600,  151200),
    ("CI/CD Add-on",     "GitHub Actions overages beyond 2000 min/mo",               672,    8064),
]
bgs = [WHITE, LGRAY, WHITE, LGRAY, WHITE]
p_colors = [GREEN, GREEN, AMBER, RED, DGRAY]
for i, (r, bg, pc) in enumerate(zip(rows, bgs, p_colors), 5):
    ws1.cell(i,2,r[0]).fill=fill(bg); ws1.cell(i,2).font=fnt(True,10,pc); ws1.cell(i,2).border=bdr(); ws1.cell(i,2).alignment=aln("left")
    ws1.cell(i,3,r[1]).fill=fill(bg); ws1.cell(i,3).font=fnt(size=9,color=DGRAY); ws1.cell(i,3).border=bdr(); ws1.cell(i,3).alignment=aln("left")
    for ci, val in enumerate(r[2:], 4):
        cl = ws1.cell(i, ci, val)
        cl.fill=fill(bg); cl.border=bdr(); cl.alignment=aln("center")
        cl.font=fnt(True,10, GREEN if val==0 else "1B2A4A")
        cl.number_format=INR_FMT
    rh(ws1, i, 20)

# SHEET 2 INITIAL COSTS
ws2 = wb.create_sheet("Initial Costs")
ws2.sheet_view.showGridLines = False
for i, w in enumerate([3, 34, 18, 16, 16, 26, 3], 1): cw(ws2, i, w)

ws2.merge_cells("B1:F1")
cl = ws2["B1"]; cl.value = "ONE-TIME / INITIAL SETUP COSTS"
cl.fill, cl.font, cl.alignment = fill(NAVY), fnt(True, 14, WHITE), aln("center")
rh(ws2, 1, 32)

hdr(ws2, 3, ["Item", "Category", "Cost (Rs)", "Frequency", "Notes"], bg=TEAL)
rh(ws2, 3, 22)

items = [
    ("Google Play Developer Account",      "Distribution",    2100, "One-time", "Lifetime access - no annual renewal"),
    ("Apple Developer Program",            "Distribution",    8316, "Yearly",   "Required for iOS App Store"),
    ("Domain Name (e.g. saferider.app)",   "Infrastructure",  1008, "Yearly",   "Via Namecheap / Google Domains"),
    ("SSL Certificate",                    "Infrastructure",     0, "Free",     "Bundled with HF Spaces & Cloudflare"),
    ("Firebase Project",                   "Backend",            0, "Free",     "No cost to create"),
    ("Hugging Face Space",                 "ML Hosting",         0, "Free",     "Docker Space creation is free"),
    ("GitHub Repository",                  "CI/CD",              0, "Free",     "Free public & private repos"),
    ("Figma / Design Assets (opt.)",       "Design",          1680, "One-time", "Figma free tier sufficient"),
]

bgs2 = [WHITE, LGRAY]*10
for i, (row, bg) in enumerate(zip(items, bgs2), 4):
    ws2.cell(i,2,row[0]).fill=fill(bg); ws2.cell(i,2).font=fnt(size=10); ws2.cell(i,2).border=bdr(); ws2.cell(i,2).alignment=aln()
    ws2.cell(i,3,row[1]).fill=fill(bg); ws2.cell(i,3).font=fnt(True,10,TEAL); ws2.cell(i,3).border=bdr(); ws2.cell(i,3).alignment=aln("center")
    cl4=ws2.cell(i,4,row[2]); cl4.fill=fill(bg); cl4.font=fnt(True,10,RED if row[2]>0 else GREEN); cl4.number_format=INR_FMT; cl4.border=bdr(); cl4.alignment=aln("center")
    ws2.cell(i,5,row[3]).fill=fill(bg); ws2.cell(i,5).font=fnt(size=10,color=DGRAY); ws2.cell(i,5).border=bdr(); ws2.cell(i,5).alignment=aln("center")
    ws2.cell(i,6,row[4]).fill=fill(bg); ws2.cell(i,6).font=fnt(size=9,italic=True,color=DGRAY); ws2.cell(i,6).border=bdr(); ws2.cell(i,6).alignment=aln("left",wrap=True)
    rh(ws2, i, 20)

tr=len(items)+4
for label, inr, bg_ in [
    ("TOTAL  Android Only",            2100,             TEAL),
    ("TOTAL  Android + iOS + Domain",  2100+8316+1008,   NAVY),
]:
    for ci, v in enumerate([label, "", inr, "", ""], 2):
        cl=ws2.cell(tr,ci,v)
        cl.fill=fill(bg_); cl.font=fnt(True,11,WHITE); cl.border=bdr(); cl.alignment=aln("center")
        if ci==4: cl.number_format=INR_FMT
    rh(ws2, tr, 24); tr+=1

# SHEET 3 MONTHLY COSTS
ws3 = wb.create_sheet("Monthly Costs")
ws3.sheet_view.showGridLines = False
for i, w in enumerate([3, 28, 18, 15, 16, 16, 16, 18, 26, 3], 1): cw(ws3, i, w)

ws3.merge_cells("B1:I1")
cl=ws3["B1"]; cl.value="MONTHLY RUNNING COSTS BY GROWTH PHASE"
cl.fill, cl.font, cl.alignment = fill(NAVY), fnt(True,14,WHITE), aln("center")
rh(ws3, 1, 32)

cols3=["Service","Category","Phase 0\nFree (Rs)","Phase 1\nStarter (Rs)","Phase 2\nGrowth (Rs)","Phase 3\nScale (Rs)","Billing","Notes"]
hdr(ws3, 3, cols3, bg=TEAL)
rh(ws3, 3, 40)
for ci in range(1,9): ws3.cell(3,ci).alignment=aln("center",wrap=True)

svcs=[
    ("Firebase Authentication",   "Auth/Backend",   0,    0,    420,  1680, "Monthly",      "Free 10K MAU. Rs0.46/verification after"),
    ("Firestore Reads & Writes",  "Database",       0,    0,    420,  2520, "Monthly",      "Free: 50K reads + 20K writes/day"),
    ("Firestore Storage",         "Database",       0,    0,     84,   840, "Monthly",      "Free 1 GB. Rs15/GB after"),
    ("Firebase Realtime DB",      "Database",       0,    0,      0,   420, "Monthly",      "Free 1 GB storage, 10 GB/mo download"),
    ("HF Spaces Backend",         "ML Hosting",     0,    0,      0,     0, "Free",         "Sleeps after 30 min inactivity"),
    ("Railway.app (always-on)",   "ML Hosting",     0,  420,    420,     0, "Monthly",      "Hobby Rs420/mo - no cold starts"),
    ("Google Cloud Run (alt.)",   "ML Hosting",     0,    0,    168,   840, "Pay/use",      "Scale-to-zero, ~Rs0.0017/vCPU-s"),
    ("GitHub Repo + Actions",     "CI/CD",          0,    0,      0,   336, "Monthly",      "2000 free min/mo. Rs0.67/min after"),
    ("Firebase App Distribution", "Beta Testing",   0,    0,      0,     0, "Free",         "Unlimited testers, always free"),
    ("Google Play Store updates", "Distribution",   0,    0,      0,     0, "Free",         "No per-update fee after Rs2100 setup"),
    ("Apple App Store",           "Distribution",   0,  693,    693,   693, "Monthly eqv.", "Rs8316/year billed annually"),
    ("Domain renewal",            "Infrastructure", 0,   84,     84,    84, "Monthly eqv.", "~Rs1008/year = Rs84/mo"),
    ("Monitoring / Sentry",       "Observability",  0,    0,      0,   840, "Monthly",      "Sentry free: 5K errors/mo; paid after"),
]

phase_sums=[0,0,0,0]
bgs3=[WHITE,LGRAY]*20
for i,(svc,bg) in enumerate(zip(svcs,bgs3),4):
    ws3.cell(i,2,svc[0]).fill=fill(bg); ws3.cell(i,2).font=fnt(True,10); ws3.cell(i,2).border=bdr(); ws3.cell(i,2).alignment=aln()
    ws3.cell(i,3,svc[1]).fill=fill(bg); ws3.cell(i,3).font=fnt(True,9,TEAL); ws3.cell(i,3).border=bdr(); ws3.cell(i,3).alignment=aln("center")
    for ci,val in enumerate(svc[2:6],4):
        cl=ws3.cell(i,ci,val)
        cl.fill=fill(bg); cl.border=bdr(); cl.alignment=aln("center")
        if val==0: cl.font=fnt(size=10,color="AAAAAA")
        else:
            cl.font=fnt(True,10,"1B2A4A"); cl.number_format=INR_FMT
            phase_sums[ci-4]+=val
    ws3.cell(i,8,svc[6]).fill=fill(bg); ws3.cell(i,8).font=fnt(size=9,color=DGRAY); ws3.cell(i,8).border=bdr(); ws3.cell(i,8).alignment=aln("center")
    ws3.cell(i,9,svc[7]).fill=fill(bg); ws3.cell(i,9).font=fnt(size=9,italic=True,color=DGRAY); ws3.cell(i,9).border=bdr(); ws3.cell(i,9).alignment=aln("left",wrap=True)
    rh(ws3,i,20)

tr3=len(svcs)+4
ws3.cell(tr3,2,"MONTHLY TOTAL (Rs)").fill=fill(NAVY); ws3.cell(tr3,2).font=fnt(True,11,WHITE); ws3.cell(tr3,2).border=bdr(); ws3.cell(tr3,2).alignment=aln("center")
ws3.cell(tr3,3,"").fill=fill(NAVY); ws3.cell(tr3,3).border=bdr()
pc_colors=[GREEN,AMBER,AMBER,RED]
for ci,(tot,pc) in enumerate(zip(phase_sums,pc_colors),4):
    cl=ws3.cell(tr3,ci,tot)
    cl.fill=fill(pc); cl.font=fnt(True,12,WHITE); cl.number_format=INR_FMT; cl.border=bdr(); cl.alignment=aln("center")
for ci in range(8,10): ws3.cell(tr3,ci,"").fill=fill(NAVY); ws3.cell(tr3,ci).border=bdr()
rh(ws3,tr3,26)

# SHEET 4 CI/CD PIPELINE
ws4=wb.create_sheet("CI-CD Pipeline")
ws4.sheet_view.showGridLines=False
for i,w in enumerate([3,26,22,22,18,20,28,3],1): cw(ws4,i,w)

ws4.merge_cells("B1:G1")
cl=ws4["B1"]; cl.value="RECOMMENDED CI/CD PIPELINE - Most Cost-Effective Setup"
cl.fill,cl.font,cl.alignment=fill(NAVY),fnt(True,14,WHITE),aln("center")
rh(ws4,1,32)

ws4.merge_cells("B2:G2")
cl=ws4["B2"]; cl.value="Stack: GitHub Free  +  GitHub Actions (2000 min/mo free)  +  Fastlane  +  Firebase App Distribution  +  HF Spaces auto-redeploy"
cl.fill,cl.font,cl.alignment=fill(TEAL),fnt(italic=True,size=10,color=WHITE),aln("center")
rh(ws4,2,18)

hdr(ws4,4,["Stage","Tool","Trigger","Monthly Cost (Rs)","Free Allowance","Purpose"],bg=TEAL)
rh(ws4,4,20)

steps=[
    ("1. Source Control",      "GitHub Free",                "Always on",          "Rs0",  "Unlimited repos",              "Store Flutter app + FastAPI backend code"),
    ("2. Code Analysis",       "flutter analyze",            "Every push",         "Rs0",  "2000 min/mo included",         "Catch type errors and lint warnings early"),
    ("3. Unit & Widget Tests", "flutter test + pytest",      "Every push",         "Rs0",  "2000 min/mo included",         "Run app tests + backend API tests"),
    ("4. Build Release AAB",   "GitHub Actions + Flutter",   "Push to main",       "Rs0",  "~8 min per build",             "Build signed Android App Bundle"),
    ("5. Beta Distribution",   "Firebase App Distribution",  "Push to main",       "Rs0",  "Unlimited testers, always",    "Auto-send APK to testers via email/link"),
    ("6. Play Store Deploy",   "Fastlane supply",            "On git release tag", "Rs0",  "Included in Rs2100 account",   "Auto-publish to internal / alpha / production"),
    ("7. Backend Redeploy",    "HF Spaces (git push)",       "Push to main",       "Rs0",  "Docker build free",            "git push triggers Docker rebuild on HF"),
    ("8. ML Model Upload",     "huggingface_hub Python SDK", "On retrain",         "Rs0",  "Free upload API",              "Push new pkl files to HF Space"),
    ("9. Error Monitoring",    "Sentry (free tier)",         "Always on",          "Rs0",  "5K errors/month free",         "Track crashes, backend errors in real time"),
    ("10. iOS Build (opt.)",   "Codemagic free tier",        "On release tag",     "Rs0",  "500 min/month free",           "macOS build without owning a Mac"),
]

bgs4=[WHITE,LGRAY]*10
for i,(step,bg) in enumerate(zip(steps,bgs4),5):
    for ci,val in enumerate(step,2):
        cl=ws4.cell(i,ci,val)
        cl.fill=fill(bg); cl.border=bdr(); cl.alignment=aln("left",wrap=True)
        if ci==2: cl.font=fnt(True,10,TEAL)
        elif ci==5:
            cl.font=fnt(True,10,GREEN if val in ("Rs0","Rs0") else AMBER)
            cl.alignment=aln("center")
        else: cl.font=fnt(size=10)
    rh(ws4,i,22)

tr4=len(steps)+5
ws4.merge_cells(start_row=tr4,start_column=2,end_row=tr4,end_column=7)
cl=ws4.cell(tr4,2,"TOTAL CI/CD COST = Rs0/month  (well within GitHub free 2000 min quota using ~500 min/mo)")
cl.fill=fill(GREEN); cl.font=fnt(True,12,WHITE); cl.alignment=aln("center")
for ci in range(2,8): ws4.cell(tr4,ci).fill=fill(GREEN); ws4.cell(tr4,ci).border=bdr()
rh(ws4,tr4,26)

r4b=tr4+2
ws4.merge_cells(f"B{r4b}:G{r4b}")
cl=ws4.cell(r4b,2,"GITHUB ACTIONS MINUTE BUDGET BREAKDOWN")
cl.fill=fill(MGRAY); cl.font=fnt(True,12,NAVY); cl.alignment=aln("center"); rh(ws4,r4b,22)

hdr(ws4,r4b+1,["Job","Tool","Duration/run","Runs/month","Minutes used","Cost"],bg=NAVY)
jobs4=[
    ("flutter analyze",      "GitHub Actions","2 min","60 pushes",  "120 min","Rs0"),
    ("flutter test",         "GitHub Actions","3 min","60 pushes",  "180 min","Rs0"),
    ("Build release AAB",    "GitHub Actions","8 min","10 releases","80 min", "Rs0"),
    ("pytest backend tests", "GitHub Actions","2 min","60 pushes",  "120 min","Rs0"),
    ("TOTAL",                "",              "",     "",           "500 min", "Rs0  (25% of free quota)"),
]
bgs4b=[WHITE,LGRAY]*10
for i,(jb,bg) in enumerate(zip(jobs4,bgs4b),r4b+2):
    bold_=jb[0]=="TOTAL"
    bg_=TEAL if bold_ else bg
    fg_=WHITE if bold_ else "1B2A4A"
    drow(ws4,i,list(jb),bg=bg_,bold=bold_,fg=fg_)
    rh(ws4,i,18)

# SHEET 5 YEARLY PROJECTION
ws5=wb.create_sheet("Yearly Projection")
ws5.sheet_view.showGridLines=False
for i,w in enumerate([3,40,18,18,3],1): cw(ws5,i,w)

ws5.merge_cells("B1:D1")
cl=ws5["B1"]; cl.value="YEARLY COST PROJECTION (All amounts in Rs)"
cl.fill,cl.font,cl.alignment=fill(NAVY),fnt(True,14,WHITE),aln("center")
rh(ws5,1,32)

hdr(ws5,3,["Expense","Year 1 (Rs)","Year 2+ (Rs)"],bg=TEAL)
rh(ws5,3,22)

yearly=[
    ("Google Play Developer (one-time)",          2100,      0),
    ("Apple Developer Program (if iOS launch)",   8316,   8316),
    ("Domain Name renewal",                       1008,   1008),
    ("Railway.app backend Rs420/mo",              5040,   5040),
    ("Firebase Auth + Firestore (avg growth)",       0,  10080),
    ("GitHub Actions (within free tier)",             0,      0),
    ("Firebase App Distribution",                    0,      0),
    ("Monitoring - Sentry free tier",                0,      0),
    ("HF Spaces backend hosting",                    0,      0),
]
y1i=sum(r[1] for r in yearly)
y2i=sum(r[2] for r in yearly)

bgs5=[WHITE,LGRAY]*10
for i,(row,bg) in enumerate(zip(yearly,bgs5),4):
    ws5.cell(i,2,row[0]).fill=fill(bg); ws5.cell(i,2).font=fnt(size=10); ws5.cell(i,2).border=bdr(); ws5.cell(i,2).alignment=aln()
    for ci,val in enumerate(row[1:],3):
        cl=ws5.cell(i,ci,val); cl.fill=fill(bg); cl.border=bdr(); cl.alignment=aln("center")
        cl.font=fnt(True,10,GREEN if val==0 else "1B2A4A")
        cl.number_format=INR_FMT
    rh(ws5,i,20)

tr5=len(yearly)+4
for label,y1,y2,bg_ in [
    ("ANNUAL TOTAL  (Android + iOS + Domain)",  y1i,       y2i,      NAVY),
    ("ANNUAL TOTAL  (Android Only - no Apple)", y1i-8316, y2i-8316, TEAL),
    ("YEAR 2+ TOTAL  (Android + iOS)",          y1i,       y2i,      GREEN),
]:
    ws5.cell(tr5,2,label).fill=fill(bg_); ws5.cell(tr5,2).font=fnt(True,11,WHITE); ws5.cell(tr5,2).border=bdr(); ws5.cell(tr5,2).alignment=aln("center")
    for ci,val in enumerate([y1,y2],3):
        cl=ws5.cell(tr5,ci,val); cl.fill=fill(bg_); cl.font=fnt(True,12,WHITE)
        cl.number_format=INR_FMT; cl.border=bdr(); cl.alignment=aln("center")
    rh(ws5,tr5,26); tr5+=1

tr5+=1
ws5.merge_cells(f"B{tr5}:D{tr5}")
cl=ws5.cell(tr5,2,"COST-SAVING TIPS"); cl.fill=fill(GREEN); cl.font=fnt(True,12,WHITE); cl.alignment=aln("center"); rh(ws5,tr5,22)

tips=[
    "Stay on HF Spaces free tier initially - 30s cold start is fine for early users",
    "Add Railway (Rs420/mo) only when instant response is needed (>100 daily active users)",
    "Firebase free tier handles ~1,000 daily active users with zero charges",
    "GitHub Actions 2,000 min/month free = ~400 full builds - enough for solo dev",
    "Use Fastlane + GitHub Actions instead of Codemagic to keep CI/CD at Rs0/month",
    "Launch Android first (Rs2,100 one-time) - defer iOS (Rs8,316/yr) until validated",
    "Use Firebase App Distribution (free) for all beta testing - no TestFlight needed",
    "HF Spaces auto-rebuilds Docker on git push - zero manual deployment steps",
]
for i,tip in enumerate(tips,tr5+1):
    ws5.merge_cells(f"B{i}:D{i}")
    cl=ws5.cell(i,2,f"  {tip}")
    cl.fill=fill(LGRAY if i%2==0 else WHITE); cl.font=fnt(size=10); cl.alignment=aln("left"); cl.border=bdr()
    rh(ws5,i,18)

path=r"d:/Rash_Driving_Pipeline2 - Copy/SafeRider_Budget_Breakdown.xlsx"
wb.save(path)
print("Saved:", path)
